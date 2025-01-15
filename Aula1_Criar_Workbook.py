from openpyxl import Workbook
from openpyxl import worksheet

# O Workbook básicamente é uma pasta de trabalho  do excel
# o Workbook já inicia com uma planilha

area_trabalho = Workbook() 

# .active é a planilha ativa da pasta de trabalho onde estamos trabalhando
ws = area_trabalho.active 

print(ws) # <Worksheet "Sheet"> pasta de trabalho com 1 planilha 
print(type(ws)) # <class 'openpyxl.worksheet.worksheet.Worksheet'>

# .title é o nome da planilha ou título da planilha

print(ws.title) # titulo da planilha padrão é sheet
print(type(ws)) # é uma class str um objeto string

# mudando o nome da planilha
nome_planilha = input('Digite o nome novo da planilha: ')

ws.title = f"{nome_planilha}"

print(ws.title) # veja que a planilha foi renomeada

# criando uma nova planilha 
ws1 = area_trabalho.create_sheet("planilha2",1) #args nome da planilha e a posição dela

print('Área de trabalho atual:')
# visualizando as planilhas que existem na minha workbook/area de trabalho no excel
print(f'{area_trabalho.sheetnames}\n') 
print(type(area_trabalho.sheetnames)) # veja que ele lista os nomes das planilhas

# com .sheet_properties podemos acessar as propriedades da planilha
ws.sheet_properties

# aqui estamos mudando a cor de fundo da célula onde fica o nome da planilha
ws.sheet_properties.tabColor = "1072BA"
print("\nCor de fundo da Sheet foi trocada")

# salvando a nossa aréa de trabalho com todo o nosso progesso
area_trabalho.save("Teste1.xlsx") # args nome do arquivo excel com estensão .xlsx
print("Arquivo excel Salvo na pasta local")