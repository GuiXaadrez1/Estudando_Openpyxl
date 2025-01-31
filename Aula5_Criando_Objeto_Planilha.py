from openpyxl import Workbook, load_workbook
import os

class Planilha(Workbook):
    
    def __init__(self):
        super().__init__() #Inicializa a classe base Workbook
        

    # Criando métodos

    # método salvar_arquivo_xlsx
    def save_sheets(self, name_archive_xlsx:str):
        self.save(f'{name_archive_xlsx}.xlsx')
        print(f'Planilha Salva como: {name_archive_xlsx}.xlsx')

    # método nomear_planilha
    def rename_sheets(self, new_name_sheet:str):
        archive_xlsx = self.active # ativando o arquivo.xlsx na planilha padrão
        archive_xlsx.title = new_name_sheet # mudando o nome dela
        print(f'Nome da Aba renomeada para: {new_name_sheet}')

    # Método para ler um arquivo .xlsx existente
    def read_sheet(self, archive_xlsx: str):
        
        # Verifica se o arquivo existe antes de tentar carregá-lo
        if os.path.exists(archive_xlsx):
            
            archive = load_workbook(archive_xlsx)  # Carrega o arquivo .xlsx
            
            sheet_names = archive.sheetnames  # Lista o nome das abas na planilha
            print(f"Arquivo '{archive_xlsx}' foi lido com sucesso!")
            print(f"Abas disponíveis no arquivo: {sheet_names}")
        else:
            print(f"Erro: O arquivo '{archive_xlsx}' não existe.")        

    # método que adiciona dados através de uma lista de dados
    def write_data_in_sheet(self,list_datas):
        
        archive = self.active
        
        for row_index, row_datas in enumerate(list_datas, start=1):  
            for column_index, item in enumerate(row_datas, start=1):  
                archive.cell(row=row_index, column=column_index, value=item)

if __name__ == "__main__":
    
    # Materializando a nossa classe Planilha em um Objeto
    p1 = Planilha()
    
    #p1.rename_sheets('p1')
    #p1.save_sheets('Aula5')
    
    #archive_xlsx = os.path.join(os.getcwd(),'Aula2.xlsx')
    
    #p1.read_sheet(archive_xlsx)

    
    lista_dados = [
        ['CODIGO', 'DESCRICAO', 'VALOR'],
        ['100','Material de Limpeza','1500'],
        ['1203','Material de Consumo', '1700'],      
        ['1245','Materia Prima', '2500']
    ]
    
    p1.write_data_in_sheet(lista_dados)
    
    p1.rename_sheets('Dados_Lista')
    p1.save_sheets('Pl_dados_ls')
    