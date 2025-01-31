from openpyxl import load_workbook
import os

# Vamos aprender a copiar planilhas

# criando caminho para o arquivo excel
arquivo_xlsx = os.path.join(os.getcwd(),'Aula2.xlsx')

# lendo o arquivo excel em uma área/pasta de trabalho Workbook

wb = load_workbook(arquivo_xlsx)

# o mp vai representar o minha planilha
# ativando a planilha ao qual vamos trabalhar
wb.active = wb['teste1']

# colcoando a planilha(ou aba) ativa dentro de um objeto(variável)

mp = wb.active = wb['teste1']

# podemos fazer assim também para ativar a planilha ao qual vamos trabalhar
#mp = wb['teste1']

'''
Quando usar cada abordagem:

Apenas acessar a aba: Use mp = wb['teste1'] se você só precisa manipular os dados da aba
sem modificar o estado da planilha (como a aba ativa).

Definir como ativa: Use wb.active = wb['teste1'] se você precisa alterar a aba ativa para
que, ao abrir o arquivo no Excel, a aba 'teste1' seja exibida por padrão.
'''

# podemos usar o método copy_worksheets para realizar essa façanha 

#mp_copy = wb.copy_worksheet(mp)

# basicamente aqui pegamos a nossa planilha ativa que iriamos trabalhar, copiamos ela e
# colocamos nesta variável mp_copy que vai representar uma planilha copiada

#mp_copy.title = "Copia do teste1"

# visualizando as planilhas que copiamos

#print(wb.sheetnames)

# Veja que todas essa planilhasa estão na masma pasta de trabalho ou área de trabalho
'''['AOB', 'planilha2', 'teste1', 'Copia do teste1']'''

# sempre salvar com extensão .xlsx
#wb.save('Aula02_Copiada.xlsx')]

# criando função para copiar planilha
def copiar_planilhas(nome_planilha):
    wb.copy_worksheet(nome_planilha)

mp_copy = copiar_planilhas(mp)

print(wb.sheetnames)

# ativando a planilha copiada
mp = wb.active = wb['teste1 Copy']
# mudando o nome
mp.title = "teste 01 copiada"

wb.save('Aula02_Copiada.xlsx')