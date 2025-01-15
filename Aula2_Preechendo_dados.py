from openpyxl import workbook
from openpyxl import worksheet
from openpyxl import load_workbook
import os

# Aula2 preechendo dados da planilha
# PRÁTICA
# vamos pegar a planilha anterior para realizar preenchimento de dados

# usando desta forma com caminhos porque acho mais seguro
arquivo_xlsx_1 = os.path.join(os.getcwd(),'Aula1.xlsx')
#print(arquivo_xlsx_1)

# Mas pode ser feito de forma direta desde que o arquivo esteja na mesma pasta:
arquivo_xlsx_2 = 'Aula1.xlsx'

# a variável wb vai ser a nossa área de trabalho(WorkBoook)
# aqui com o load_workbook estamos lendo o arquivo.xlsx
wb = load_workbook(arquivo_xlsx_1)

# Ativando planilha ao qual vamos trabalhar
wb.active = wb['AOB']

# Mudando de planilha
wb2 = wb.active = wb['planilha2']
print(f'planilha foi mudada para: {wb2}')

# criando uma nova planilha worksheet de forma diferente
print("\nUma nova planilha foi criada:")
teste = wb.create_sheet('teste1')
print(teste)

# Listando todas as nossas planilhas criadas
print('\nListando Todas as nossas Sheets')
print(wb.sheetnames)

wb.active = teste
print(f"Mudando para a planilha: {teste}")

# Preenchedo planilha através do nome da célula:
teste['A1'] = "Primeiro"

# Preenchedo através da função .cell()
teste.cell(row=1,column=1,value=100) # args = linha [index], coluna [index], value = valor 

# preenchendo dinâmicamente as células com for in for loop
# veja que coloquei um valor fixo, mas pode ser mudado

'''
for row in range(1,20):
    for column in range(1,20):
        teste.cell(row = row, column = column, value = 10)
'''


# Outro exemplo de preencimento dinâmico através de uma lista

lista_dados = [
    ['CODIGO', 'DESCRICAO', 'VALOR'],
    ['100','Material de Limpeza','1500'],
    ['1203','Material de Consumo', '1700'],      
    ['1245','Materia Prima', '2500']
    ]

print(type(lista_dados))


# Realizando o mesmo for in for loop só que com o valores da lista de dados

for row_index, row_dados in enumerate(lista_dados, start=1):  
    for column_index, item in enumerate(row_dados, start=1):  
        teste.cell(row=row_index, column=column_index, value=item)


wb.save('Aula2.xlsx')