from openpyxl import load_workbook  # Método para ler planilhas Excel existentes
import os

# Criando o caminho para a planilha "Aula2.xlsx"
arquivo_xlsx = os.path.join(os.getcwd(), 'Aula2.xlsx')

# print(arquivo_xlsx)  # Caso queira verificar o caminho completo do arquivo

# Carregando o arquivo Excel na memória
wb = load_workbook(arquivo_xlsx)  # Carrega a planilha para manipulação

# Exibindo os nomes das abas presentes na planilha
print(wb.sheetnames)  # Retorna uma lista com os títulos das abas da planilha

print()  # Apenas uma linha em branco para organização
for i in wb.sheetnames:
    print(i)  # Exibe cada aba da lista individualmente
print()

# Ativando a aba específica onde vamos trabalhar
ws = wb['teste1']  # Especificamos a aba desejada pelo nome

# Usando `.values` para acessar os dados presentes na aba especificada
# `.values` retorna um iterador onde cada linha da planilha é uma tupla

# Loop simples para exibir todas as linhas da aba como tuplas
for row in ws.values:
    print(row)

print()  # Linha em branco para separar os resultados

# Loop duplo para exibir os valores de cada célula, linha por linha
for row in ws.values:  # Primeiro loop percorre cada linha (como tupla)
    for value in row:  # Segundo loop percorre cada valor dentro da linha
        print(value)
print()

# Resumo:
# - O primeiro loop (`for row in ws.values`) itera sobre as linhas da planilha.
#   Cada linha é representada como uma tupla contendo os valores das células.
# - O segundo loop (`for value in row`) percorre os valores dentro de cada tupla,
#   permitindo acessar os dados célula por célula.
# - Isso é útil para acessar e manipular os dados dentro de uma aba específica da planilha.

# Outra forma de acessarmos os dados é usando o método iter_rows
# O métodos podemos passar parâmetros representando linhas e colunas, formando células
# ele nos permitir vizualizar ou não os dados com o parâmetros values_only = True
# isso tira a referência da célula e retorna o valor dela. Veja!

for row in ws.iter_rows(min_row = 1, max_col = 3, max_row = 4, values_only = True): 
    print(row)

# Basicamente faz a mesma coisa que isso aqui
'''
for row in ws.values:
    print(row)
'''
# Porém, podemos definir o tamanho da nossa planilha e a quantidade de dados que queremos
# retornar dela